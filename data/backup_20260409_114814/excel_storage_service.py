# backend/services/excel_storage_service.py
import os
import pandas as pd
from typing import Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import quote_sheetname
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from PIL import Image as PILImage
from io import BytesIO
import logging

from backend.schemas.table_schemas import ExcelSaveConfig

logger = logging.getLogger(__name__)


class ExcelStorageService:
    def __init__(self):
        pass

    def save_dataframe(self, df: pd.DataFrame, excel_path: str, sheet_name: str,
                       map_name: str, image_data: Optional[Union[str, bytes]] = None,
                       config: Optional[ExcelSaveConfig] = None) -> bool:
        """
        将DataFrame保存到Excel，支持图片插入和目录管理

        Args:
            df: 要保存的DataFrame
            excel_path: Excel文件路径
            sheet_name: 工作表名称
            map_name: 目录中显示的表名
            image_data: 图片数据（文件路径、bytes或PIL.Image）
            config: Excel保存配置

        Returns:
            bool: 保存是否成功
        """
        if config is None:
            config = ExcelSaveConfig()

        try:
            if not excel_path.endswith('.xlsx'):
                raise ValueError("文件路径必须以.xlsx结尾")

            # 加载或创建工作簿
            wb = self._load_or_create_workbook(excel_path)

            # 处理工作表
            ws = self._prepare_worksheet(wb, sheet_name, config.mode)

            # 写入数据
            self._write_dataframe_to_sheet(ws, df)

            # 处理目录
            self._update_catalog(wb, sheet_name, map_name)

            # 插入图片
            if image_data is not None:
                self._insert_image_to_sheet(ws, image_data, config.anchor_cell, config.width_px)

            # 保存工作簿
            wb.save(excel_path)

            logger.info(f'已按"{config.mode}"模式写入"{sheet_name}"表 → {excel_path}')
            if image_data:
                logger.info(f'  并插入图片 @ {config.anchor_cell}')
            logger.info(f'  目录已更新：{sheet_name} -> {map_name}')

            return True

        except Exception as e:
            logger.error(f"保存到Excel失败: {e}")
            return False

    def _load_or_create_workbook(self, excel_path: str) -> Workbook:
        """加载或创建工作簿"""
        try:
            if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                return load_workbook(excel_path)
            else:
                wb = Workbook()
                # 清理默认工作表
                if len(wb.sheetnames) > 0:
                    default_sheet = wb.active.title
                    if default_sheet != '目录':
                        del wb[default_sheet]
                return wb
        except Exception as e:
            logger.warning(f"文件损坏，创建新文件: {e}")
            wb = Workbook()
            if len(wb.sheetnames) > 0:
                del wb[wb.active.title]
            return wb

    def _prepare_worksheet(self, wb: Workbook, sheet_name: str, mode: str):
        """准备工作表"""
        if mode == 'overwrite' and sheet_name in wb.sheetnames:
            del wb[sheet_name]
        return wb.create_sheet(sheet_name)

    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame):
        """将DataFrame写入工作表"""
        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # 写入数据行
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _update_catalog(self, wb: Workbook, sheet_name: str, map_name: str):
        """更新目录"""
        catalog = '目录'
        if catalog not in wb.sheetnames:
            cat_ws = wb.create_sheet(catalog, 0)
            cat_ws.append(['sheet_name', 'table_name'])
        else:
            cat_ws = wb[catalog]

        # 生成超链接
        quoted_sheet = quote_sheetname(sheet_name)
        loc = f"#{quoted_sheet}!A1"

        # 更新或新增目录行
        updated = False
        for row in cat_ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if row[0].value == sheet_name:
                row[1].value = map_name
                row[1].hyperlink = Hyperlink(ref=row[1].coordinate, location=loc)
                row[1].font = Font(underline='single', color='0563C1')
                updated = True
                break

        if not updated:
            new_row = cat_ws.max_row + 1
            cat_ws.cell(row=new_row, column=1, value=sheet_name)
            cell = cat_ws.cell(row=new_row, column=2)
            cell.value = map_name
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=loc)
            cell.font = Font(underline='single', color='0563C1')

    def _insert_image_to_sheet(self, ws, image_data: Union[str, bytes], anchor_cell: str, width_px: int):
        """插入图片到工作表"""
        try:
            # 处理不同类型的image_data
            if isinstance(image_data, bytes):
                pil_img = PILImage.open(BytesIO(image_data))
            elif isinstance(image_data, str) and os.path.isfile(image_data):
                pil_img = PILImage.open(image_data)
            elif isinstance(image_data, PILImage.Image):
                pil_img = image_data
            else:
                raise TypeError("image_data必须是文件路径、bytes或PIL.Image对象")

            # 调整图片尺寸
            if width_px:
                ratio = pil_img.height / pil_img.width
                pil_img = pil_img.resize((width_px, int(width_px * ratio)), PILImage.LANCZOS)

            # 保存到临时流并插入Excel
            tmp = BytesIO()
            pil_img.save(tmp, format='PNG')
            tmp.seek(0)
            xl_img = XLImage(tmp)
            xl_img.anchor = anchor_cell
            ws.add_image(xl_img)

        except Exception as e:
            logger.error(f"图片插入失败: {e}")

    def read_excel_sheet(self, excel_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """读取Excel工作表中的数据"""
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel文件不存在: {excel_path}")
                return None

            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            logger.info(f"成功读取工作表: {sheet_name}")
            return df

        except Exception as e:
            logger.error(f"读取Excel失败: {e}")
            return None

    def get_sheet_names(self, excel_path: str) -> list:
        """获取Excel文件中的所有工作表名称"""
        try:
            if not os.path.exists(excel_path):
                return []

            wb = load_workbook(excel_path)
            return wb.sheetnames

        except Exception as e:
            logger.error(f"获取工作表名称失败: {e}")
            return []

    def create_new_excel(self, excel_path: str) -> bool:
        """创建新的Excel文件"""
        try:
            wb = Workbook()
            # 删除默认工作表
            if len(wb.sheetnames) > 0:
                del wb[wb.active.title]
            wb.save(excel_path)
            logger.info(f"创建新的Excel文件: {excel_path}")
            return True

        except Exception as e:
            logger.error(f"创建Excel文件失败: {e}")
            return False