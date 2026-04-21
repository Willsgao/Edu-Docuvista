# -*- coding:utf-8 -*-
"""
Test table rebuild logic
Batch generate Excel from multiple OCR and LLM data pairs
"""

import json
import gzip
import os
import sys
import io
import re
from pprint import pprint

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Add backend to path
# backend_path = r"E:\Datas\base_pros\DocuVista"
backend_path = r"F:\wills\codes\DocuVista"
sys.path.insert(0, backend_path)

# ==================== Configuration ====================
OCR_DIR = fr"{backend_path}\test_codes\test_pic_data\ocr"
LLM_DIR = fr"{backend_path}\test_codes\test_pic_data\llm"

OUTPUT_EXCEL_PATH = fr"{backend_path}\test_codes\test_output_batch_new.xlsx"


def find_matching_pairs(ocr_dir, llm_dir):
    """Scan directories and find matching OCR-LLM file pairs by prefix"""
    ocr_files = {}
    llm_files = {}
    
    for f in os.listdir(ocr_dir):
        if f.endswith('.json.gz'):
            match = re.match(r'^(\d+)_', f)
            if match:
                prefix = match.group(1)
                ocr_files[prefix] = f
    
    for f in os.listdir(llm_dir):
        if f.endswith('.json.gz'):
            match = re.match(r'^(\d+)_', f)
            if match:
                prefix = match.group(1)
                llm_files[prefix] = f
    
    common_prefixes = set(ocr_files.keys()) & set(llm_files.keys())
    pairs = []
    for prefix in sorted(common_prefixes):
        pairs.append({
            'prefix': prefix,
            'ocr_file': os.path.join(ocr_dir, ocr_files[prefix]),
            'llm_file': os.path.join(llm_dir, llm_files[prefix])
        })
    
    return pairs


def load_json_gz(file_path):
    """Load json.gz file"""
    print(f"Loading file: {file_path}")
    with gzip.open(file_path, 'rt', encoding='utf-8') as f:
        data = json.load(f)
    print(f"Loaded successfully")
    return data


def save_excel(all_tables_data, metadata_dict, output_path):
    """
    使用 openpyxl 直接写入（与 step9_save_to_excel_optimized 逻辑一致），
    每个 prefix 的所有表格写入同一个 Excel，多个 sheet。

    metadata_dict: { prefix: metadata_list }
    all_tables_data: [(tables_data, table_names, prefix), ...]
    """
    if not all_tables_data:
        print("No table data")
        return False

    try:
        from openpyxl import Workbook
        from pathlib import Path

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        for tables_data, table_names, prefix in all_tables_data:
            metadata_list = metadata_dict.get(prefix, [None] * len(tables_data))
            for idx, (table, name) in enumerate(zip(tables_data, table_names)):
                sheet_name = f"{prefix}_{name}"[:31]
                ws = wb.create_sheet(title=sheet_name)

                filtered_table = [row[:] for row in table]

                # 删除 row_marker=0 的第2行（与 step9 一致）
                if len(table) >= 2:
                    second_row = table[1]
                    last_val = second_row[-1] if second_row else None
                    if str(last_val).strip() in ['0']:
                        del filtered_table[1]

                # 写入数据
                for r, row in enumerate(filtered_table, 1):
                    if r == 1:
                        ws.cell(row=r, column=1, value="项目0")
                    else:
                        ws.cell(row=r, column=1, value="")
                    for c, val in enumerate(row, 2):
                        ws.cell(row=r, column=c, value=val)

                # 元数据
                if metadata_list and idx < len(metadata_list):
                    metadata = metadata_list[idx]
                    if metadata:
                        data_row_count = len(filtered_table)
                        metadata_start_row = data_row_count + 2
                        valid_keys = ["bankname", "currency", "report_period",
                                      "unit", "table_name", "ocr_table_id", "entity"]
                        field_mapping = {
                            "bankname": "bank_name",
                            "currency": "default_currency",
                            "report_period": "default_report_period",
                            "unit": "default_unit",
                            "table_name": "original_table_name",
                            "ocr_table_id": "ocr_table_id",
                            "entity": "entity"
                        }
                        ws.cell(row=metadata_start_row, column=1, value="")
                        for row_offset, key in enumerate(valid_keys, 1):
                            value = metadata.get(field_mapping.get(key, key), "")
                            ws.cell(row=metadata_start_row + row_offset, column=1,
                                    value=f"{key}:{value}")
                        ws.cell(row=metadata_start_row + len(valid_keys) + 1, column=1, value="")

        wb.save(output_path)
        wb.close()
        print(f"Excel saved: {output_path}")
        return True
    except Exception as e:
        print(f"Save Excel failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def print_data_structure(data, name, indent=0):
    """Recursively print data structure"""
    prefix = "  " * indent
    
    if isinstance(data, dict):
        print(f"{prefix}{name}: dict with {len(data)} keys")
        for key, value in data.items():
            if isinstance(value, (dict, list)) and value:
                print_data_structure(value, f"{key}", indent + 1)
            else:
                value_type = type(value).__name__
                if isinstance(value, str) and len(value) > 50:
                    value_repr = repr(value[:50] + "...")
                elif isinstance(value, list) and len(value) > 10:
                    value_repr = f"list with {len(value)} items"
                else:
                    value_repr = repr(value)[:100]
                print(f"{prefix}  {key}: {value_type} = {value_repr}")
    elif isinstance(data, list):
        print(f"{prefix}{name}: list with {len(data)} items")
        if data:
            if isinstance(data[0], dict):
                print(f"{prefix}  [0] item keys: {list(data[0].keys())}")
                for i, item in enumerate(data[:3]):
                    print(f"{prefix}  --- Item {i} ---")
                    print_data_structure(item, "", indent + 1)
                if len(data) > 3:
                    print(f"{prefix}  ... and {len(data) - 3} more items")
            elif isinstance(data[0], list):
                print(f"{prefix}  [0] is list with {len(data[0])} items")
                print(f"{prefix}  First 3 rows:")
                for i, row in enumerate(data[:3]):
                    print(f"{prefix}    Row {i}: {row[:5]}{'...' if len(row) > 5 else ''}")
            else:
                print(f"{prefix}  First item: {repr(data[0])[:100]}")
    else:
        print(f"{prefix}{name}: {type(data).__name__} = {repr(data)[:100]}")


def compare_data_structures(ocr_data, llm_data):
    """Compare and print differences between OCR and LLM data structures"""
    print("\n" + "=" * 60)
    print("数据结构对比分析")
    print("=" * 60)
    
    print("\n>>> OCR 数据顶层结构:")
    print_data_structure(ocr_data, "OCR")
    print("111111111111111111111111111:")
    print(type(ocr_data))
    pprint(ocr_data)
    
    print("\n>>> LLM 数据顶层结构:")
    print_data_structure(llm_data, "LLM")
    print("222222222222222222222222222:")
    pprint(llm_data)
    
    print("\n>>> 异同点分析:")
    print("-" * 40)
    
    ocr_keys = set(ocr_data.keys()) if isinstance(ocr_data, dict) else set()
    llm_keys = set(llm_data.keys()) if isinstance(llm_data, dict) else set()
    
    common_keys = ocr_keys & llm_keys
    ocr_only = ocr_keys - llm_keys
    llm_only = llm_keys - ocr_keys
    
    print(f"共同keys: {common_keys if common_keys else '无'}")
    print(f"OCR独有keys: {ocr_only if ocr_only else '无'}")
    print(f"LLM独有keys: {llm_only if llm_only else '无'}")
    
    print("\n>>> 详细对比:")
    print("-" * 40)
    
    if 'tables_result' in ocr_data and 'tables' in llm_data:
        ocr_tables = ocr_data.get('tables_result', [])
        llm_tables = llm_data.get('tables', [])
        
        print(f"\nOCR tables_result: {len(ocr_tables)} tables")
        print(f"LLM tables: {len(llm_tables)} tables")
        
        # Print OCR table content in grid format
        print("\n" + "=" * 60)
        print("OCR 表格详细内容 (Table 1 - 有数据的表格)")
        print("=" * 60)
        tbl = ocr_tables[1]  # Table 1 has data
        if 'body' in tbl:
            body = tbl['body']
            print(f"总单元格数: {len(body)}")
            
            # Find max row and col to reconstruct the table
            max_row = max(cell['row_end'] for cell in body)
            max_col = max(cell['col_end'] for cell in body)
            print(f"表格维度: {max_row} rows x {max_col} cols")
            
            # Build 2D grid
            grid = [['' for _ in range(max_col)] for _ in range(max_row)]
            for cell in body:
                r = cell['row_start']
                c = cell['col_start']
                grid[r][c] = cell['words']
            
            print("\n表格内容 (按行):")
            for i, row in enumerate(grid):
                print(f"Row {i}: {row}")
        
        # Print LLM headers
        print("\n" + "=" * 60)
        print("LLM 表格结构 (Table 0)")
        print("=" * 60)
        tbl = llm_tables[0]
        print(f"表格名称: {tbl.get('name')}")
        print(f"对应OCR表格索引: {tbl.get('ocr_tables')}")
        
        headers = tbl.get('headers', {})
        print(f"\n列标题(cols): {headers.get('cols')}")
        print(f"行标题(rows): {headers.get('rows')}")
        print(f"货币: {tbl.get('default_currency')}")
        print(f"单位: {tbl.get('default_unit')}")
        print(f"报表期间: {tbl.get('default_report_period')}")
        
        # 分析差异
        print("\n" + "=" * 60)
        print("问题分析")
        print("=" * 60)
        print(f"- OCR识别到 {max_col} 列数据")
        print(f"- LLM识别到 {len(headers.get('cols', []))} 列表头")
        print(f"- OCR第1列内容(第一列): {[row[0] for row in grid]}")
        print(f"- LLM行标题: {headers.get('rows')}")


def is_date_text(text):
    """判断是否是日期相关的文本"""
    if not text or not isinstance(text, str):
        return False
    text = str(text).strip()
    date_keywords = ['年', '月', '日', '截至', '止']
    return any(kw in text for kw in date_keywords) and ('/' in text or len(text) < 20)


def has_numeric_data(row, exclude_first_col=True):
    """判断行是否有数值数据"""
    if not row or len(row) == 0:
        return False
    start_idx = 1 if exclude_first_col else 0
    for col_idx in range(start_idx, len(row)):
        cell = row[col_idx]
        if cell and cell not in [None, '']:
            cell_str = str(cell).strip()
            if cell_str and not is_date_text(cell_str):
                return True
    return False


def check_first_round_issues(table_data, llm_rows):
    """
    检测第一轮是否存在严重问题：
    1. 数据行数与LLM行标题数不匹配
    2. 数据行的第一列缺少完整层级标题（不包含">>"）
    """
    if not table_data or not llm_rows:
        return False, "无数据"
    
    # 统计有数值的数据行数量
    data_row_count = 0
    for row_idx, row in enumerate(table_data):
        if row_idx == 0:
            continue
        if has_numeric_data(row):
            data_row_count += 1
    
    llm_rows_count = len(llm_rows)
    
    # 检查1：数据行数与LLM行标题数不匹配
    if data_row_count != llm_rows_count:
        return True, f"数据行数({data_row_count}) != LLM行标题数({llm_rows_count})"
    
    # 检查2：数据行的第一列是否缺少完整层级标题
    missing_headers = 0
    for row_idx, row in enumerate(table_data):
        if row_idx == 0:
            continue
        if has_numeric_data(row) and len(row) > 0:
            first_col = row[0]
            if not first_col or '>>' not in str(first_col):
                missing_headers += 1
    
    if missing_headers > 0:
        return True, f"有{missing_headers}行缺少完整层级表头"
    
    return False, "无问题"


def _extract_year_from_header(header_str):
    """从列标题中提取年份，如 '2025年1-6月' -> 2025"""
    if not header_str or not isinstance(header_str, str):
        return None
    m = re.search(r'(\d{4})', header_str)
    if m:
        return int(m.group(1))
    return None


def _extract_year_from_data(cell):
    """从数据单元格中提取年份，如 '2025年1-6月' -> 2025"""
    if not cell or not isinstance(cell, str):
        return None
    m = re.search(r'(\d{4})', cell)
    if m:
        return int(m.group(1))
    return None


def fix_column_issues(table_data, llm_cols, llm_rows, table_idx, table_name, ocr_cells=None):
    """
    检测并修复列错位问题。

    核心问题：step5 误删年份列后，header行和data行列数不一致。
    例如 header=[None,2025,2024,2023,None] (5列) 但 data=[label,2025,2024,2023] (4列)，
    pandas 读取时会错位：2024数据 -> 2025列，2023数据 -> 2024列，None -> 2023列。

    修复原则：
    1. 确保 header 行和 data 行的列数一致（都等于 LLM cols 数）
    2. 年份数据列按 LLM cols 的年份顺序重对齐
    3. 当列数多于 LLM 列数时，利用 OCR span 信息删除冗余列
    """
    if not table_data or len(table_data) == 0 or not llm_cols:
        return table_data, False

    import copy
    working = copy.deepcopy(table_data)
    original_col_count = len(working[0])
    llm_col_count = len(llm_cols)

    print(f"\n  [列检测 表格{table_idx}] {table_name}")
    print(f"    原始列数: {original_col_count}, LLM列数: {llm_col_count}")

    # Step 1: 从 LLM cols 提取年份顺序
    llm_year_to_pos = {}  # year -> pos in llm_cols
    llm_year_list = []
    for pos, col in enumerate(llm_cols):
        year = _extract_year_from_header(col)
        if year:
            llm_year_to_pos[year] = pos
            llm_year_list.append(year)

    if not llm_year_list:
        print(f"    LLM中无年份列，仅补齐列数")
        return _fix_column_count_only(working, llm_cols, table_idx, table_name, ocr_cells=ocr_cells)

    print(f"    LLM年份顺序: {llm_year_list}")

    # [修复-20260404] Step 1.5: 当列数多于 LLM 列数时，先删除冗余列
    if original_col_count > llm_col_count:
        excess_cols = original_col_count - llm_col_count
        print(f"    [冗余列检测] 列数多于LLM {excess_cols}列，尝试删除...")

        # 空列检测：从 row1 开始，找到完全为空的列
        empty_col_indices = []
        for col_idx in range(original_col_count):
            is_empty = True
            for row_idx in range(1, min(len(working), 20)):
                if col_idx < len(working[row_idx]):
                    cell = working[row_idx][col_idx]
                    if cell is not None and str(cell).strip() != '':
                        # 含中文 = 有效列
                        if any('\u4e00' <= ch <= '\u9fff' for ch in str(cell)):
                            is_empty = False
                            break
                        # 纯数字 = 有效数据列
                        clean = str(cell).replace(',', '').replace(' ', '')
                        if clean.replace('.', '').replace('-', '').replace('(', '').replace(')', '').replace(' ', '').isdigit():
                            is_empty = False
                            break
                        # 其他非空字符串
                        is_empty = False
                        break
            if is_empty:
                empty_col_indices.append(col_idx)

        print(f"    [冗余列检测] 空列: {empty_col_indices} (共{len(empty_col_indices)}个)")

        cols_to_delete = []
        if len(empty_col_indices) >= excess_cols:
            # 空列足够，优先删除右侧
            cols_to_delete = sorted(empty_col_indices, reverse=True)[:excess_cols]
            print(f"    [冗余列检测] 空列足够，删除右侧: {cols_to_delete}")
        elif ocr_cells and len(empty_col_indices) < excess_cols:
            # 空列不够，用 OCR span 信息找更多冗余列
            still_need = excess_cols - len(empty_col_indices)
            span_redundant = []
            for col_idx in range(original_col_count):
                if col_idx in empty_col_indices:
                    continue
                # 检查数据行(row2+)是否全空
                data_rows_empty = True
                for row_idx in range(2, min(len(working), 20)):
                    if col_idx < len(working[row_idx]):
                        cell = working[row_idx][col_idx]
                        if cell is not None and str(cell).strip() != '':
                            data_rows_empty = False
                            break
                if not data_rows_empty:
                    continue
                # 检查是否被 OCR span>1 的 cell 覆盖
                covered_by_span = False
                for cell in ocr_cells:
                    cs = cell['col_start']
                    ce = cell['col_end']
                    if ce - cs > 1 and cs <= col_idx < ce:
                        covered_by_span = True
                        break
                if covered_by_span:
                    span_redundant.append(col_idx)
                    print(f"    [冗余列检测] col {col_idx}: 数据行全空，被OCR span覆盖 -> 冗余列")

            if len(span_redundant) >= still_need:
                extra_delete = sorted(span_redundant, reverse=True)[:still_need]
                cols_to_delete = sorted(empty_col_indices + extra_delete, reverse=True)
                print(f"    [冗余列检测] 补充span冗余列: {extra_delete}，最终删除: {cols_to_delete}")
            else:
                print(f"    [冗余列检测] span冗余列不足（{len(span_redundant)} < {still_need}），跳过删除")

        # 执行删除（从右向左）
        if cols_to_delete:
            for i in range(len(working)):
                for col_idx in cols_to_delete:
                    if col_idx < len(working[i]):
                        del working[i][col_idx]
            print(f"    [冗余列检测] 删除完成，列数 {original_col_count} -> {len(working[0])}")
            original_col_count = len(working[0])

    # Step 2: 从数据行检测每列的年份
    # 注意：header行(第0行)可能年份值不全，要从data行(第1-5行)检测
    data_col_years = {}
    for col_idx in range(len(working[0])):
        for row_idx in range(1, min(len(working), 6)):
            if col_idx >= len(working[row_idx]):
                break
            year = _extract_year_from_data(working[row_idx][col_idx])
            if year:
                data_col_years[col_idx] = year
                break

    print(f"    数据中年份列: {data_col_years}")

    # Step 3: 如果 header 行和 data 行列数不一致，先对齐列数
    data_col_count = len(working[1]) if len(working) > 1 else original_col_count
    header_col_count = original_col_count

    if header_col_count != data_col_count:
        print(f"    !! Header列数({header_col_count}) != Data列数({data_col_count})，先对齐列数")
        # 目标：header列数 = data列数
        # 如果 header 比 data 少，在末尾补 None
        # 如果 header 比 data 多...（这种情况较少）
        if header_col_count < data_col_count:
            needed = data_col_count - header_col_count
            for row in working:
                for _ in range(needed):
                    row.append(None)
            header_col_count = data_col_count
            original_col_count = data_col_count
            print(f"    -> Header补齐到{header_col_count}列")

    # Step 4: 检查年份列错位
    # 对于每列 col，数据中年份=data_col_years[col]，LLM期望年份在pos=llm_year_to_pos[year]
    # 如果 col != pos，说明该列的数据应该属于 LLM 中 pos 位置的年份
    needs_swap = []
    for col_idx, data_year in data_col_years.items():
        if data_year not in llm_year_to_pos:
            continue
        expected_pos = llm_year_to_pos[data_year]
        if col_idx != expected_pos:
            needs_swap.append((col_idx, expected_pos, data_year))
            print(f"    !! 年份{data_year}：数据在col{col_idx}，应在col{expected_pos}")

    if not needs_swap:
        print(f"    PASS: 年份列无需交换")
        # 确保列数够（补 None）
        return _fix_column_count_only(working, llm_cols, table_idx, table_name, ocr_cells=ocr_cells)

    # Step 5: 执行交换（只交换 header 行和数据行，不碰行标题列）
    # 需要找到 LLM 中"第0列"的定义
    # 如果 llm_cols[0] 不包含年份 → 第0列是行标题（保留原位）
    # 如果 llm_cols[0] 包含年份 → 所有列都是数据列
    if llm_year_list and llm_year_to_pos.get(llm_year_list[0], -1) != 0:
        # llm_cols[0] 不是年份，第0列是行标题
        row_label_col = 0
        is_row_label_col = True
    else:
        row_label_col = -1  # 没有行标题列
        is_row_label_col = False

    print(f"    行标题列: col{row_label_col} (is_row_label={is_row_label_col})")

    # 对每一对需要交换的 (data_col, expected_pos)
    # 在 header 行交换：
    # 在 data 行交换：需要确保不破坏行标题列
    for data_col, expected_pos, year in needs_swap:
        # 跳过行标题列（如果有的话）
        if is_row_label_col and (data_col == row_label_col or expected_pos == row_label_col):
            print(f"    跳过：涉及行标题列 col{data_col}<->col{expected_pos}")
            continue

        # 交换 header 行
        working[0][data_col], working[0][expected_pos] = working[0][expected_pos], working[0][data_col]
        # 交换所有 data 行
        for row_i in range(1, len(working)):
            if data_col < len(working[row_i]) and expected_pos < len(working[row_i]):
                working[row_i][data_col], working[row_i][expected_pos] = working[row_i][expected_pos], working[row_i][data_col]

        print(f"    交换: col{data_col}<->col{expected_pos} (年份{year})")

    # Step 6: 补齐缺失列（确保 header 和 data 列数 = LLM cols 数）
    final_col_count = len(working[0])
    target_cols = llm_col_count

    if final_col_count < target_cols:
        missing = target_cols - final_col_count
        print(f"    补齐缺失的 {missing} 列为 None")
        for row in working:
            for _ in range(missing):
                row.append(None)

    print(f"    PASS: 列修复完成，最终列数={len(working[0])}")
    return working, True


def _fix_column_count_only(table_data, llm_cols, table_idx, table_name, ocr_cells=None):
    """只处理列数不匹配（不处理列偏移）"""
    if not table_data or len(table_data) == 0:
        return table_data, False

    actual_cols = len(table_data[0])
    has_row_header = (actual_cols == len(llm_cols) + 1)
    expected_data_cols = len(llm_cols) if llm_cols else 0

    if has_row_header:
        target_cols = expected_data_cols + 1
    else:
        target_cols = expected_data_cols

    fixed = False

    if actual_cols < target_cols and target_cols > 0:
        missing = target_cols - actual_cols
        print(f"    ⚠️ 列数少于预期，缺失 {missing} 列 → 用 None 补齐")
        pad_cols = [None] * missing
        for row_idx in range(len(table_data)):
            table_data[row_idx] = list(table_data[row_idx]) + pad_cols
        fixed = True
        print(f"    ✅ 补齐完成，新列数: {len(table_data[0])}")

    elif has_row_header and actual_cols > target_cols:
        extra = actual_cols - target_cols
        # [修复-20260404] 尝试检测空列并删除
        empty_cols = []
        for col_idx in range(actual_cols):
            is_empty = True
            for row_idx in range(1, min(len(table_data), 20)):
                if col_idx < len(table_data[row_idx]):
                    cell = table_data[row_idx][col_idx]
                    if cell is not None and str(cell).strip() != '':
                        is_empty = False
                        break
            if is_empty:
                empty_cols.append(col_idx)
        if len(empty_cols) >= extra:
            delete_cols = sorted(empty_cols, reverse=True)[:extra]
            print(f"    ⚠️ 列数多于预期，删除空列: {delete_cols}")
            for i in range(len(table_data)):
                for col_idx in delete_cols:
                    if col_idx < len(table_data[i]):
                        del table_data[i][col_idx]
            fixed = True
            print(f"    ✅ 删除完成，新列数: {len(table_data[0])}")
        else:
            print(f"    ⚠️ 列数多于预期 {extra} 列，空列仅 {len(empty_cols)} 个 → 不自动删除")

    elif not has_row_header and llm_cols and actual_cols == len(llm_cols):
        print(f"    ✓ 列数与 LLM 匹配，无列问题")

    return table_data, fixed


def fix_row_issues(table_data, llm_rows, llm_cols, table_idx, table_name):
    """
    检测并修复行标题问题：用 LLM rows 替换有数值的数据行的第一列。

    策略（与原逻辑一致，但匹配关系正确）：
    1. 统计有数值的数据行数量
    2. 如果数据行数 == LLM 行标题数：用对应 LLM 行标题逐行替换
    3. 如果不匹配：仍然尝试替换（兜底逻辑）
    """
    if not table_data or len(table_data) == 0:
        return table_data, False

    # 统计有数值的数据行（跳过第0行，即列标题行）
    data_row_count = 0
    for row_idx in range(1, len(table_data)):
        if has_numeric_data(table_data[row_idx]):
            data_row_count += 1

    llm_rows_count = len(llm_rows) if llm_rows else 0

    print(f"\n  [行检测 表格{table_idx}] {table_name}")
    print(f"    有数值的数据行数: {data_row_count}, LLM行标题数: {llm_rows_count}")

    # 判断是否需要修复：数据行缺少 ">>" 层级标题
    missing_header_count = 0
    for row_idx in range(1, len(table_data)):
        row = table_data[row_idx]
        if has_numeric_data(row) and len(row) > 0:
            first_cell = row[0]
            if not first_cell or '>>' not in str(first_cell):
                missing_header_count += 1

    has_issue = (missing_header_count > 0) or (data_row_count != llm_rows_count)

    if not has_issue:
        print(f"    ✓ 无行标题问题")
        return table_data, False

    print(f"    ⚠️ 发现行标题问题：{missing_header_count}行缺少'>>'层级标题")
    print(f"    → 执行二次处理，替换有数值的数据行")

    new_table = []
    llm_idx = 0

    for row_idx, row in enumerate(table_data):
        if not row or len(row) == 0:
            new_table.append(list(row) if row else [])
            continue

        if row_idx == 0:
            # 保留列标题行
            new_table.append(list(row))
            continue

        if has_numeric_data(row):
            # 关键修复：如果第一列是日期，说明这行是年份行
            # （被step7删除最左空列后，年份从col0被推到了col1）
            # 不能执行 row[1:]，否则会丢失年份数据
            if is_date_text(row[0]):
                new_table.append(list(row))
                continue

            if llm_idx < llm_rows_count:
                llm_row_header = llm_rows[llm_idx]
                new_row = [llm_row_header] + list(row[1:]) if len(row) > 1 else [llm_row_header]
                llm_idx += 1
                print(f"    Row {row_idx}: 替换 LLM[{llm_idx-1}] = {llm_row_header}")
            else:
                new_row = list(row)
        else:
            new_row = list(row)

        new_table.append(new_row)

    print(f"    ✅ 修复完成：共替换 {llm_idx} 行")
    return new_table, True


def detect_and_fix_header_issues(tables_data, table_names, llm_result, ocr_result=None):
    """
    二次处理：检测并修复第一轮遗留的表头问题。

    修复策略：
    1. 正确匹配：tables_data[i] <-> llm_tables[i]（不再永远只用 llm_tables[0]）
    2. 先修列：用 LLM cols 检测/补齐缺失的列，利用 OCR span 信息删除冗余列
    3. 再修行：用 LLM rows 修复缺失的行标题

    仅在第一轮存在问题时才执行修复。
    """
    if not tables_data or not llm_result:
        return tables_data, table_names

    # 与第一轮 (_process_single_table_to_memory) 保持一致，优先从 tables_structure 读取
    llm_tables = llm_result.get('tables_structure', {}).get('tables', [])
    if not llm_tables:
        llm_tables = llm_result.get('tables', [])
    if not llm_tables:
        return tables_data, table_names

    print("\n" + "=" * 60)
    print("二次处理：检测并修复表头问题（修正匹配逻辑）")
    print("=" * 60)
    print(f"tables_data 数量: {len(tables_data)}, llm_tables 数量: {len(llm_tables)}")

    fixed_tables = []
    for i, table_data in enumerate(tables_data):
        if not table_data:
            fixed_tables.append(table_data)
            continue

        # =============================================
        # 关键修复：正确获取对应的 LLM 表格
        # =============================================
        if i < len(llm_tables):
            llm_table = llm_tables[i]
            llm_headers = llm_table.get('headers', {})
            llm_cols = llm_headers.get('cols', [])
            llm_rows = llm_headers.get('rows', [])
        else:
            # 如果 llm_tables 不够，用第0个（兜底）
            print(f"\n  ⚠️ 表格 {i} 超出 llm_tables 范围（len={len(llm_tables)}），使用 llm_tables[0] 兜底")
            llm_headers = llm_tables[0].get('headers', {})
            llm_cols = llm_headers.get('cols', [])
            llm_rows = llm_headers.get('rows', [])

        table_name = table_names[i] if i < len(table_names) else 'unknown'

        # [修复-20260404] 获取该表格对应的 OCR cells（用于 span 冗余列检测）
        table_ocr_cells = None
        if ocr_result and i < len(llm_tables):
            llm_table = llm_tables[i]
            ocr_table_indices = llm_table.get('ocr_tables', [])
            if ocr_table_indices and 'tables_result' in ocr_result:
                ocr_idx = ocr_table_indices[0]
                if ocr_idx < len(ocr_result['tables_result']):
                    table_ocr_cells = ocr_result['tables_result'][ocr_idx].get('body', [])

        print(f"\n{'─' * 50}")
        print(f"表格 {i}: {table_name}")
        print(f"  LLM cols 数量: {len(llm_cols)}, LLM rows 数量: {len(llm_rows)}")

        # [DEBUG] 打印原始表格前5行
        import copy
        print(f"  [DEBUG] 原始表格前5行:")
        for ri, row in enumerate(table_data[:5]):
            print(f"    Row {ri} ({len(row)} cols): {row}")
        print(f"  [DEBUG] 原始表格总列数: {len(table_data[0]) if table_data else 0}")

        working_table = copy.deepcopy(table_data)

        # Step 1: 修复列问题
        working_table, col_fixed = fix_column_issues(
            working_table, llm_cols, llm_rows, i, table_name, ocr_cells=table_ocr_cells
        )

        # Step 2: 修复行问题
        working_table, row_fixed = fix_row_issues(
            working_table, llm_rows, llm_cols, i, table_name
        )

        if not col_fixed and not row_fixed:
            print(f"  ✓ 无需修复")
        else:
            print(f"  ✅ 修复完成")

        fixed_tables.append(working_table)

    return fixed_tables, table_names


def main():
    print("=" * 60)
    print("批量处理测试")
    print("=" * 60)
    
    try:
        import pandas
    except ImportError:
        print("ERROR: pandas not installed")
        return
    
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed")
        return
    
    pairs = find_matching_pairs(OCR_DIR, LLM_DIR)
    print(f"\n发现 {len(pairs)} 对文件:")
    for p in pairs:
        print(f"  - {p['ocr_file']} <-> {p['llm_file']}")
    
    all_results = []
    metadata_dict = {}  # {prefix: metadata_list}
    
    for pair in pairs:
        prefix = pair['prefix']
        ocr_file = pair['ocr_file']
        llm_file = pair['llm_file']
        
        print(f"\n{'='*50}")
        print(f"处理中: {prefix}")
        print(f"{'='*50}")
        
        print(f"\nLoading OCR: {os.path.basename(ocr_file)}")
        with gzip.open(ocr_file, 'rt', encoding='utf-8') as f:
            ocr_result = json.load(f)


        print("11111111111111ocr_result1111111111111111")
        pprint(ocr_result)
        
        print(f"Loading LLM: {os.path.basename(llm_file)}")
        with gzip.open(llm_file, 'rt', encoding='utf-8') as f:
            llm_result = json.load(f)

        print("11111111111111llm_result1111111111111111")
        print(llm_result)
        
        try:
            from backend.core.table_processor.table_rebuilder import TableReconstructor
            
            rebuilder = TableReconstructor()
            
            tables_data, table_names, metadata_list = rebuilder.process_all_tables_to_memory(
                ocr_result=ocr_result,
                llm_result=llm_result,
                image_path=f"test_{prefix}.jpg",
                bank_name="Test Bank"
            )
            
            tables_data, table_names = detect_and_fix_header_issues(tables_data, table_names, llm_result, ocr_result)
            metadata_dict[prefix] = metadata_list
            
            all_results.append((tables_data, table_names, prefix))
            print(f"  ✅ 成功: {len(tables_data)} 个表格")
            
        except Exception as e:
            print(f"  ❌ 失败: {e}")
            import traceback
            traceback.print_exc()
    
    if all_results:
        print(f"\n{'='*60}")
        print("保存Excel...")
        print(f"{'='*60}")
        save_excel(all_results, metadata_dict, OUTPUT_EXCEL_PATH)
    
    print("\n" + "=" * 60)
    print("批量测试完成")
    print("=" * 60)


if __name__ == "__main__":
    main()
